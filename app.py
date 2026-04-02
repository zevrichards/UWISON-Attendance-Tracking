import os
import csv
import io
from datetime import datetime
from functools import wraps
from contextlib import contextmanager

from flask import (Flask, render_template, request, redirect, url_for,
                   session, flash, jsonify, Response)
from werkzeug.security import generate_password_hash, check_password_hash
import psycopg2
import psycopg2.extras
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "change-me-in-production")

DATABASE_URL  = os.environ.get("DATABASE_URL", "")
DEDUP_SECONDS = int(os.environ.get("DEDUP_SECONDS", "30"))
# Build version — derived from git at startup
import subprocess as _sp
def _get_version():
    try:
        sha = _sp.check_output(
            ["git", "rev-parse", "--short", "HEAD"],
            stderr=_sp.DEVNULL
        ).decode().strip()
        count = _sp.check_output(
            ["git", "rev-list", "--count", "HEAD"],
            stderr=_sp.DEVNULL
        ).decode().strip()
        return f"{count}.{sha}"
    except Exception:
        return "dev"
BUILD_VERSION = _get_version()
# Build version — derived from git at startup
import subprocess as _sp
def _get_version():
    try:
        sha = _sp.check_output(
            ["git", "rev-parse", "--short", "HEAD"],
            stderr=_sp.DEVNULL
        ).decode().strip()
        count = _sp.check_output(
            ["git", "rev-list", "--count", "HEAD"],
            stderr=_sp.DEVNULL
        ).decode().strip()
        return f"{count}.{sha}"
    except Exception:
        return "dev"
BUILD_VERSION = _get_version()

# Render gives postgres:// URLs; psycopg2 needs postgresql://
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)


# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------

@contextmanager
def get_db():
    """Yield a psycopg2 connection with RealDictCursor; auto-commit on success."""
    conn = psycopg2.connect(DATABASE_URL,
                            cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def init_db():
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id       SERIAL PRIMARY KEY,
                username TEXT   UNIQUE NOT NULL,
                pw_hash  TEXT   NOT NULL,
                role     TEXT   NOT NULL DEFAULT 'staff'
            );

            CREATE TABLE IF NOT EXISTS courses (
                id         SERIAL PRIMARY KEY,
                code       TEXT NOT NULL UNIQUE,
                name       TEXT NOT NULL,
                created_at TEXT NOT NULL
            );
            -- Add unique constraint if table already exists without it
            DO $$
            BEGIN
                IF NOT EXISTS (
                    SELECT 1 FROM pg_constraint
                    WHERE conname = 'courses_code_key'
                ) THEN
                    ALTER TABLE courses ADD CONSTRAINT courses_code_key UNIQUE (code);
                END IF;
            END$$;

            CREATE TABLE IF NOT EXISTS students (
                id         SERIAL PRIMARY KEY,
                course_id  INTEGER NOT NULL REFERENCES courses(id),
                student_id TEXT NOT NULL,
                name       TEXT NOT NULL,
                UNIQUE(course_id, student_id)
            );

            CREATE TABLE IF NOT EXISTS sessions (
                id         SERIAL PRIMARY KEY,
                course_id  INTEGER NOT NULL REFERENCES courses(id),
                label      TEXT NOT NULL,
                date       TEXT NOT NULL,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS attendance (
                id         SERIAL PRIMARY KEY,
                session_id INTEGER NOT NULL REFERENCES sessions(id),
                student_id TEXT NOT NULL,
                time_in    TEXT,
                time_out   TEXT,
                last_scan  TEXT,
                UNIQUE(session_id, student_id)
            );
            """)


# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


def create_user(username, password, role="staff"):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO users (username, pw_hash, role) VALUES (%s, %s, %s)",
                (username.lower(), generate_password_hash(password), role)
            )


# ---------------------------------------------------------------------------
# Routes — Auth
# ---------------------------------------------------------------------------

@app.route("/", methods=["GET"])
@login_required
def index():
    return redirect(url_for("courses"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"].lower().strip()
        password = request.form["password"]
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT * FROM users WHERE username=%s", (username,))
                user = cur.fetchone()
        if user and check_password_hash(user["pw_hash"], password):
            session["user_id"] = user["id"]
            session["username"] = user["username"]
            session["role"] = user["role"]
            return redirect(url_for("courses"))
        flash("Invalid username or password.", "error")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ---------------------------------------------------------------------------
# Routes — Courses
# ---------------------------------------------------------------------------

@app.route("/courses")
@login_required
def courses():
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT c.*,
                       COUNT(DISTINCT s.id)  AS student_count,
                       COUNT(DISTINCT ss.id) AS session_count
                FROM courses c
                LEFT JOIN students s  ON s.course_id  = c.id
                LEFT JOIN sessions ss ON ss.course_id = c.id
                GROUP BY c.id
                ORDER BY c.created_at DESC
            """)
            courses_list = cur.fetchall()
    return render_template("courses.html", courses=courses_list)


@app.route("/courses/new", methods=["GET", "POST"])
@login_required
def new_course():
    """Upload a Banner class list to create/update a course and roster in one step."""
    if request.method == "POST":
        f = request.files.get("roster")
        if not f or not f.filename:
            flash("Please select a file.", "error")
            return redirect(request.url)

        fname = f.filename.lower()
        rows  = []

        if fname.endswith(".xlsx") or fname.endswith(".xls"):
            import openpyxl as _xl
            wb       = _xl.load_workbook(f, read_only=True, data_only=True)
            ws       = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                flash("Spreadsheet appears to be empty.", "error")
                return redirect(request.url)
            headers = [str(h).strip().lower() if h else "" for h in all_rows[0]]
            for raw in all_rows[1:]:
                row = {headers[i]: (str(v).strip() if v is not None else "")
                       for i, v in enumerate(raw)}
                rows.append(row)
        elif fname.endswith(".csv"):
            stream = io.StringIO(f.stream.read().decode("utf-8-sig"))
            reader = csv.DictReader(stream)
            rows   = [{k.strip().lower(): (v or "").strip()
                       for k, v in row.items()} for row in reader]
        else:
            flash("Please upload a Banner .xlsx export or a .csv file.", "error")
            return redirect(request.url)

        if not rows:
            flash("File appears to be empty.", "error")
            return redirect(request.url)

        sample = rows[0]
        keys   = list(sample.keys())

        # Extract course code and name from Banner columns
        if "crs_code" not in keys or "course" not in keys:
            flash("Could not detect Banner columns (crs_code, course). Please use a Banner xlsx export.", "error")
            return redirect(request.url)

        code = sample.get("crs_code", "").strip().upper()
        name = sample.get("course",   "").strip()

        if not code or not name:
            flash("Course code or name is blank in the file.", "error")
            return redirect(request.url)

        # Detect student ID and name columns
        id_col = "id" if "id" in keys else next((k for k in keys if "id" in k), None)
        has_split_names = "last_name" in keys and "first_name" in keys
        name_col = None if has_split_names else next(
            (k for k in keys if "name" in k and "last" not in k
             and "first" not in k and "middle" not in k), None)

        if not id_col or (not has_split_names and not name_col):
            flash(f"Could not detect student ID/name columns. Found: {keys}.", "error")
            return redirect(request.url)

        # Upsert course — update name if code already exists
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """INSERT INTO courses (code, name, created_at)
                       VALUES (%s, %s, %s)
                       ON CONFLICT (code)
                       DO UPDATE SET name = EXCLUDED.name
                       RETURNING id""",
                    (code, name, datetime.utcnow().isoformat())
                )
                course_id = cur.fetchone()["id"]

                added = updated = skipped = 0
                for row in rows:
                    sid = row.get(id_col, "").strip().lstrip("'")
                    if has_split_names:
                        sname = f"{row.get('first_name','').strip()} {row.get('last_name','').strip()}".strip()
                    else:
                        sname = row.get(name_col, "").strip()

                    if not sid or not sname or sid.lower() in ("none", "nan", ""):
                        skipped += 1
                        continue

                    cur.execute(
                        """INSERT INTO students (course_id, student_id, name)
                           VALUES (%s, %s, %s)
                           ON CONFLICT (course_id, student_id)
                           DO UPDATE SET name = EXCLUDED.name
                           RETURNING (xmax = 0) AS inserted""",
                        (course_id, sid, sname)
                    )
                    result = cur.fetchone()
                    if result and result["inserted"]:
                        added += 1
                    else:
                        updated += 1

        parts = [f"{added} student(s) added"]
        if updated:
            parts.append(f"{updated} updated")
        if skipped:
            parts.append(f"{skipped} skipped")
        flash(f"Course {code} — {name} ready. {', '.join(parts)}.", "success")
        return redirect(url_for("course_detail", course_id=course_id))

    return render_template("new_course.html")


@app.route("/courses/<int:course_id>")
@login_required
def course_detail(course_id):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM courses WHERE id=%s", (course_id,))
            course = cur.fetchone()
            if not course:
                flash("Course not found.", "error")
                return redirect(url_for("courses"))
            cur.execute("""
                SELECT ss.*, COUNT(a.id) AS scan_count
                FROM sessions ss
                LEFT JOIN attendance a ON a.session_id = ss.id
                WHERE ss.course_id=%s
                GROUP BY ss.id
                ORDER BY ss.date DESC
            """, (course_id,))
            sessions_list = cur.fetchall()
            cur.execute("SELECT COUNT(*) AS c FROM students WHERE course_id=%s", (course_id,))
            student_count = cur.fetchone()["c"]
    return render_template("course_detail.html", course=course,
                           sessions=sessions_list, student_count=student_count)


# ---------------------------------------------------------------------------
# Routes — Roster upload
# ---------------------------------------------------------------------------

@app.route("/courses/<int:course_id>/upload-roster", methods=["GET", "POST"])
@login_required
def upload_roster(course_id):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM courses WHERE id=%s", (course_id,))
            course = cur.fetchone()
    if not course:
        return redirect(url_for("courses"))

    if request.method == "POST":
        f = request.files.get("roster")
        if not f or not f.filename:
            flash("Please select a file.", "error")
            return redirect(request.url)

        fname = f.filename.lower()
        rows  = []

        # ── Parse file ────────────────────────────────────────────────
        if fname.endswith(".xlsx") or fname.endswith(".xls"):
            # Read xlsx with openpyxl
            import openpyxl as _xl
            wb = _xl.load_workbook(f, read_only=True, data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                flash("Spreadsheet appears to be empty.", "error")
                return redirect(request.url)
            # Header row — normalise keys
            headers = [str(h).strip().lower() if h else "" for h in all_rows[0]]
            for raw in all_rows[1:]:
                row = {headers[i]: (str(v).strip() if v is not None else "")
                       for i, v in enumerate(raw)}
                rows.append(row)

        elif fname.endswith(".csv"):
            stream = io.StringIO(f.stream.read().decode("utf-8-sig"))
            reader = csv.DictReader(stream)
            rows   = [{k.strip().lower(): (v or "").strip()
                       for k, v in row.items()} for row in reader]
        else:
            flash("Please upload a Banner .xlsx export or a .csv file.", "error")
            return redirect(request.url)

        if not rows:
            flash("File appears to be empty.", "error")
            return redirect(request.url)

        sample = rows[0]
        keys   = list(sample.keys())

        # ── Detect ID column ──────────────────────────────────────────
        # Banner export uses exact column "id"; fallback to any col containing "id"
        if "id" in keys:
            id_col = "id"
        else:
            id_col = next((k for k in keys if "id" in k), None)

        # ── Detect name columns ───────────────────────────────────────
        # Banner has separate last_name / first_name columns.
        # Simple rosters may have a single "name" column.
        has_split_names = "last_name" in keys and "first_name" in keys
        if not has_split_names:
            # Try generic single-column detection
            name_col = next((k for k in keys
                             if "name" in k and "last" not in k and "first" not in k
                             and "middle" not in k), None)
        else:
            name_col = None  # will build from first_name + last_name

        if not id_col or (not has_split_names and not name_col):
            flash(
                f"Could not detect required columns. Found: {keys}. "
                "Expected: 'id' and 'last_name'+'first_name' (Banner export) "
                "or 'id' and 'name' (simple CSV).",
                "error"
            )
            return redirect(request.url)

        # ── Insert students ───────────────────────────────────────────
        added = updated = skipped = 0
        with get_db() as conn:
            with conn.cursor() as cur:
                for row in rows:
                    sid = row.get(id_col, "").strip()
                    # Strip any leading apostrophe Excel sometimes adds to IDs
                    sid = sid.lstrip("'")

                    if has_split_names:
                        first  = row.get("first_name", "").strip()
                        last   = row.get("last_name",  "").strip()
                        name   = f"{first} {last}".strip()
                    else:
                        name = row.get(name_col, "").strip()

                    if not sid or not name or sid.lower() in ("none", "nan", ""):
                        skipped += 1
                        continue

                    # ON CONFLICT: if student already exists update their name
                    # (handles re-upload after name corrections)
                    cur.execute(
                        """INSERT INTO students (course_id, student_id, name)
                           VALUES (%s, %s, %s)
                           ON CONFLICT (course_id, student_id)
                           DO UPDATE SET name = EXCLUDED.name
                           RETURNING (xmax = 0) AS inserted""",
                        (course_id, sid, name)
                    )
                    result = cur.fetchone()
                    if result and result["inserted"]:
                        added += 1
                    else:
                        updated += 1

        parts = [f"{added} student(s) added"]
        if updated:
            parts.append(f"{updated} updated")
        if skipped:
            parts.append(f"{skipped} skipped (blank rows)")
        flash(f"Roster uploaded: {', '.join(parts)}.", "success")
        return redirect(url_for("course_detail", course_id=course_id))

    return render_template("upload_roster.html", course=course)


# ---------------------------------------------------------------------------
# Routes — Sessions
# ---------------------------------------------------------------------------

@app.route("/courses/<int:course_id>/sessions/new", methods=["POST"])
@login_required
def new_session(course_id):
    label = request.form["label"].strip()
    date  = request.form["date"]
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """INSERT INTO sessions (course_id, label, date, created_at)
                   VALUES (%s,%s,%s,%s) RETURNING id""",
                (course_id, label, date, datetime.utcnow().isoformat())
            )
            session_id = cur.fetchone()["id"]
            cur.execute("SELECT student_id FROM students WHERE course_id=%s", (course_id,))
            students = cur.fetchall()
            for s in students:
                cur.execute(
                    """INSERT INTO attendance (session_id, student_id)
                       VALUES (%s,%s)
                       ON CONFLICT (session_id, student_id) DO NOTHING""",
                    (session_id, s["student_id"])
                )
    return redirect(url_for("scan_session", session_id=session_id))


@app.route("/sessions/<int:session_id>")
@login_required
def scan_session(session_id):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT ss.*, c.code, c.name AS course_name
                FROM sessions ss
                JOIN courses c ON c.id = ss.course_id
                WHERE ss.id=%s
            """, (session_id,))
            sess = cur.fetchone()
            if not sess:
                flash("Session not found.", "error")
                return redirect(url_for("courses"))
            cur.execute("""
                SELECT a.*, s.name
                FROM attendance a
                JOIN students s ON s.student_id=a.student_id AND s.course_id=%s
                WHERE a.session_id=%s
                ORDER BY s.name
            """, (sess["course_id"], session_id))
            records = cur.fetchall()
    present = sum(1 for r in records if r["time_in"])
    total   = len(records)
    return render_template("scan_session.html", sess=sess, records=records,
                           present=present, total=total,
                           dedup_seconds=DEDUP_SECONDS)


# ---------------------------------------------------------------------------
# API — Scan endpoint
# ---------------------------------------------------------------------------

@app.route("/api/scan", methods=["POST"])
@login_required
def api_scan():
    data       = request.get_json()
    session_id = data.get("session_id")
    student_id = str(data.get("student_id", "")).strip()

    if not session_id or not student_id:
        return jsonify(status="error", message="Missing data"), 400

    now     = datetime.utcnow()
    now_str = now.isoformat()

    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM sessions WHERE id=%s", (session_id,))
            sess = cur.fetchone()
            if not sess:
                return jsonify(status="error", message="Session not found"), 404

            cur.execute(
                "SELECT * FROM students WHERE course_id=%s AND student_id=%s",
                (sess["course_id"], student_id)
            )
            student = cur.fetchone()
            if not student:
                return jsonify(status="error",
                               message=f"ID {student_id} not on roster"), 404

            cur.execute(
                """INSERT INTO attendance (session_id, student_id)
                   VALUES (%s,%s)
                   ON CONFLICT (session_id, student_id) DO NOTHING""",
                (session_id, student_id)
            )
            cur.execute(
                "SELECT * FROM attendance WHERE session_id=%s AND student_id=%s",
                (session_id, student_id)
            )
            rec = cur.fetchone()

            # Dedup check
            if rec["last_scan"]:
                last = datetime.fromisoformat(rec["last_scan"])
                diff = (now - last).total_seconds()
                if diff < DEDUP_SECONDS:
                    return jsonify(
                        status="duplicate",
                        message=f"Already scanned {int(diff)}s ago — ignored",
                        name=student["name"]
                    )

            if not rec["time_in"]:
                cur.execute(
                    """UPDATE attendance SET time_in=%s, last_scan=%s
                       WHERE session_id=%s AND student_id=%s""",
                    (now_str, now_str, session_id, student_id)
                )
                action = "time_in"
            elif not rec["time_out"]:
                cur.execute(
                    """UPDATE attendance SET time_out=%s, last_scan=%s
                       WHERE session_id=%s AND student_id=%s""",
                    (now_str, now_str, session_id, student_id)
                )
                action = "time_out"
            else:
                cur.execute(
                    "UPDATE attendance SET last_scan=%s WHERE session_id=%s AND student_id=%s",
                    (now_str, session_id, student_id)
                )
                return jsonify(
                    status="done",
                    message=f"{student['name']} already fully logged",
                    name=student["name"]
                )

    local_time = now.strftime("%H:%M:%S")
    return jsonify(
        status="ok",
        action=action,
        name=student["name"],
        student_id=student_id,
        time=local_time,
        message=f"{'Time in' if action == 'time_in' else 'Time out'} logged for {student['name']}"
    )


# ---------------------------------------------------------------------------
# API — Live records for a session
# ---------------------------------------------------------------------------

@app.route("/api/session/<int:session_id>/records")
@login_required
def api_session_records(session_id):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM sessions WHERE id=%s", (session_id,))
            sess = cur.fetchone()
            cur.execute("""
                SELECT a.*, s.name
                FROM attendance a
                JOIN students s ON s.student_id=a.student_id AND s.course_id=%s
                WHERE a.session_id=%s
                ORDER BY s.name
            """, (sess["course_id"], session_id))
            records = cur.fetchall()
    return jsonify([dict(r) for r in records])


# ---------------------------------------------------------------------------
# Routes — Semester report
# ---------------------------------------------------------------------------

@app.route("/courses/<int:course_id>/report")
@login_required
def semester_report(course_id):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM courses WHERE id=%s", (course_id,))
            course = cur.fetchone()
            cur.execute(
                "SELECT * FROM sessions WHERE course_id=%s ORDER BY date", (course_id,))
            sessions_list = cur.fetchall()
            cur.execute(
                "SELECT * FROM students WHERE course_id=%s ORDER BY name", (course_id,))
            students = cur.fetchall()

            report = []
            for st in students:
                row = {"student_id": st["student_id"], "name": st["name"], "sessions": {}}
                attended = 0
                for sess in sessions_list:
                    cur.execute(
                        "SELECT * FROM attendance WHERE session_id=%s AND student_id=%s",
                        (sess["id"], st["student_id"])
                    )
                    rec     = cur.fetchone()
                    present = bool(rec and rec["time_in"])
                    row["sessions"][sess["id"]] = {
                        "present":  present,
                        "time_in":  rec["time_in"]  if rec else None,
                        "time_out": rec["time_out"] if rec else None,
                    }
                    if present:
                        attended += 1
                total_sessions = len(sessions_list)
                pct = round((attended / total_sessions * 100) if total_sessions else 0, 1)
                row.update(attended=attended, total=total_sessions,
                           pct=pct, at_risk=pct < 75)
                report.append(row)

    return render_template("report.html", course=course,
                           sessions=sessions_list, report=report)


# ---------------------------------------------------------------------------
# Export — CSV per session
# ---------------------------------------------------------------------------

@app.route("/sessions/<int:session_id>/export/csv")
@login_required
def export_session_csv(session_id):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT ss.*, c.code
                FROM sessions ss JOIN courses c ON c.id=ss.course_id
                WHERE ss.id=%s
            """, (session_id,))
            sess = cur.fetchone()
            cur.execute("""
                SELECT a.*, s.name
                FROM attendance a
                JOIN students s ON s.student_id=a.student_id AND s.course_id=%s
                WHERE a.session_id=%s ORDER BY s.name
            """, (sess["course_id"], session_id))
            records = cur.fetchall()

    def fmt(t):
        if not t:
            return ""
        try:
            return datetime.fromisoformat(t).strftime("%H:%M:%S")
        except Exception:
            return t

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Student ID", "Name", "Time In", "Time Out", "Present"])
    for r in records:
        writer.writerow([r["student_id"], r["name"],
                         fmt(r["time_in"]), fmt(r["time_out"]),
                         "Yes" if r["time_in"] else "No"])

    filename = f"{sess['code']}_{sess['label'].replace(' ','_')}_{sess['date']}.csv"
    return Response(output.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment; filename={filename}"})


# ---------------------------------------------------------------------------
# Export — Excel semester report
# ---------------------------------------------------------------------------

@app.route("/courses/<int:course_id>/export/excel")
@login_required
def export_semester_excel(course_id):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM courses WHERE id=%s", (course_id,))
            course = cur.fetchone()
            cur.execute(
                "SELECT * FROM sessions WHERE course_id=%s ORDER BY date", (course_id,))
            sessions_list = cur.fetchall()
            cur.execute(
                "SELECT * FROM students WHERE course_id=%s ORDER BY name", (course_id,))
            students = cur.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    risk_fill   = PatternFill("solid", fgColor="FFCCCC")
    ok_fill     = PatternFill("solid", fgColor="CCFFCC")
    header_font = Font(color="FFFFFF", bold=True)

    headers = (["Student ID", "Name"] +
               [f"{s['label']} ({s['date']})" for s in sessions_list] +
               ["Sessions Attended", "Total Sessions", "Attendance %", "Status"])
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, st in enumerate(students, 2):
        ws.cell(row=row_idx, column=1, value=st["student_id"])
        ws.cell(row=row_idx, column=2, value=st["name"])
        attended = 0
        with get_db() as conn:
            with conn.cursor() as cur:
                for col_idx, sess in enumerate(sessions_list, 3):
                    cur.execute(
                        "SELECT * FROM attendance WHERE session_id=%s AND student_id=%s",
                        (sess["id"], st["student_id"])
                    )
                    rec     = cur.fetchone()
                    present = bool(rec and rec["time_in"])
                    cell       = ws.cell(row=row_idx, column=col_idx,
                                         value="P" if present else "A")
                    cell.fill      = ok_fill if present else risk_fill
                    cell.alignment = Alignment(horizontal="center")
                    if present:
                        attended += 1

        total  = len(sessions_list)
        pct    = round(attended / total * 100, 1) if total else 0
        status = "AT RISK" if pct < 75 else "OK"
        base   = 3 + len(sessions_list)
        ws.cell(row=row_idx, column=base,     value=attended)
        ws.cell(row=row_idx, column=base + 1, value=total)
        pct_cell    = ws.cell(row=row_idx, column=base + 2, value=f"{pct}%")
        status_cell = ws.cell(row=row_idx, column=base + 3, value=status)
        if pct < 75:
            pct_cell.fill    = risk_fill
            status_cell.fill = risk_fill
            status_cell.font = Font(bold=True, color="CC0000")

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{course['code']}_attendance_report.xlsx"
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ---------------------------------------------------------------------------
# Admin — user management
# ---------------------------------------------------------------------------

@app.route("/admin/users", methods=["GET", "POST"])
@login_required
def admin_users():
    if session.get("role") != "admin":
        flash("Admin access required.", "error")
        return redirect(url_for("courses"))

    if request.method == "POST":
        action = request.form.get("action")
        if action == "create":
            username = request.form["username"].strip().lower()
            password = request.form["password"]
            role     = request.form.get("role", "staff")
            try:
                create_user(username, password, role)
                flash(f"User '{username}' created.", "success")
            except Exception as e:
                flash(f"Error: {e}", "error")
        elif action == "delete":
            uid = request.form["user_id"]
            if str(uid) == str(session["user_id"]):
                flash("Cannot delete yourself.", "error")
            else:
                with get_db() as conn:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM users WHERE id=%s", (uid,))
                flash("User deleted.", "success")

    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id, username, role FROM users ORDER BY username")
            users = cur.fetchall()
    return render_template("admin_users.html", users=users)


# ---------------------------------------------------------------------------
# Startup
# ---------------------------------------------------------------------------

def bootstrap():
    init_db()
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) AS c FROM users")
            if cur.fetchone()["c"] == 0:
                admin_pw = os.environ.get("ADMIN_PASSWORD", "changeme123")
                create_user("admin", admin_pw, "admin")
                print(f"[boot] Created default admin. Password: {admin_pw}")


bootstrap()

@app.context_processor
def inject_version():
    return dict(build_version=BUILD_VERSION)

if __name__ == "__main__":
    app.run(debug=True, port=5000)


# ---------------------------------------------------------------------------
# Camera test page
# ---------------------------------------------------------------------------

@app.route("/camera-test")
@login_required
def camera_test():
    return render_template("camera_test.html")


# ---------------------------------------------------------------------------
# Backup — download full ZIP (admin only)
# ---------------------------------------------------------------------------

@app.route("/admin/backup")
@login_required
def backup():
    if session.get("role") != "admin":
        flash("Admin access required.", "error")
        return redirect(url_for("courses"))

    import zipfile

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        with get_db() as conn:
            with conn.cursor() as cur:

                # courses
                cur.execute("SELECT * FROM courses ORDER BY id")
                out = io.StringIO()
                w = csv.writer(out)
                w.writerow(["id","code","name","created_at"])
                for r in cur.fetchall():
                    w.writerow([r["id"],r["code"],r["name"],r["created_at"]])
                zf.writestr("courses.csv", out.getvalue())

                # students
                cur.execute("SELECT * FROM students ORDER BY course_id, id")
                out = io.StringIO()
                w = csv.writer(out)
                w.writerow(["id","course_id","student_id","name"])
                for r in cur.fetchall():
                    w.writerow([r["id"],r["course_id"],r["student_id"],r["name"]])
                zf.writestr("students.csv", out.getvalue())

                # sessions
                cur.execute("SELECT * FROM sessions ORDER BY course_id, id")
                out = io.StringIO()
                w = csv.writer(out)
                w.writerow(["id","course_id","label","date","created_at"])
                for r in cur.fetchall():
                    w.writerow([r["id"],r["course_id"],r["label"],r["date"],r["created_at"]])
                zf.writestr("sessions.csv", out.getvalue())

                # attendance — enriched with names/labels for readability
                cur.execute("""
                    SELECT a.id, a.session_id,
                           c.code  AS course_code,
                           ss.label AS session_label,
                           ss.date  AS session_date,
                           a.student_id,
                           s.name   AS student_name,
                           a.time_in, a.time_out
                    FROM attendance a
                    JOIN sessions ss ON ss.id = a.session_id
                    JOIN courses  c  ON c.id  = ss.course_id
                    JOIN students s  ON s.student_id = a.student_id
                                    AND s.course_id  = ss.course_id
                    ORDER BY ss.course_id, a.session_id, s.name
                """)
                out = io.StringIO()
                w = csv.writer(out)
                w.writerow(["id","session_id","course_code","session_label",
                            "session_date","student_id","student_name",
                            "time_in","time_out"])
                for r in cur.fetchall():
                    w.writerow([r["id"],r["session_id"],r["course_code"],
                                r["session_label"],r["session_date"],
                                r["student_id"],r["student_name"],
                                r["time_in"] or "",r["time_out"] or ""])
                zf.writestr("attendance.csv", out.getvalue())

                # users (no password hashes exported)
                cur.execute("SELECT id, username, role FROM users ORDER BY id")
                out = io.StringIO()
                w = csv.writer(out)
                w.writerow(["id","username","role"])
                for r in cur.fetchall():
                    w.writerow([r["id"],r["username"],r["role"]])
                zf.writestr("users.csv", out.getvalue())

    buf.seek(0)
    ts       = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"labattend_backup_{ts}.zip"
    return Response(
        buf.getvalue(),
        mimetype="application/zip",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
